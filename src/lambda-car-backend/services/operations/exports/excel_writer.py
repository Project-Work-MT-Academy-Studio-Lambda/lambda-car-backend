from io import BytesIO

from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from .rows.user_export_row import UserExportRow
from .rows.car_export_row import CarExportRow
from .rows.commit_export_row import CommitExportRow
from .rows.trip_export_row import TripExportRow
from .rows.refueling_export_row import RefuelingExportRow


class ExcelExportWriter:
    def write(
        self,
        users: list[UserExportRow],
        cars: list[CarExportRow],
        commits: list[CommitExportRow],
        trips: list[TripExportRow],
        refuelings: list[RefuelingExportRow],
    ) -> bytes:
        workbook = Workbook()

        default_sheet = workbook.active
        workbook.remove(default_sheet)

        self._write_users(workbook, users)
        self._write_cars(workbook, cars)
        self._write_commits(workbook, commits)
        self._write_trips(workbook, trips)
        self._write_refuelings(workbook, refuelings)

        output = BytesIO()
        workbook.save(output)
        output.seek(0)

        return output.getvalue()

    def _write_users(self, workbook: Workbook, rows: list[UserExportRow]) -> None:
        sheet = workbook.create_sheet("Users")
        self._append_header(sheet, ["Nome", "Email", "Ruolo"])

        for row in rows:
            sheet.append([
                row.name,
                row.email,
                row.role,
            ])

        self._autosize_columns(sheet)

    def _write_cars(self, workbook: Workbook, rows: list[CarExportRow]) -> None:
        sheet = workbook.create_sheet("Cars")
        self._append_header(sheet, [
            "Targa",
            "Modello",
            "Km totali",
            "Km tagliando",
            "Km gomme",
            "Tipo carburante",
            "Livello carburante",
            "Carta carburante",
            "CO2 per km",
        ])

        for row in rows:
            sheet.append([
                row.plate,
                row.model,
                row.km_total,
                row.km_servicing,
                row.km_wheels,
                row.fuel_type,
                row.fuel_level,
                row.fuel_card,
                row.co2_per_km,
            ])

        self._autosize_columns(sheet)

    def _write_commits(self, workbook: Workbook, rows: list[CommitExportRow]) -> None:
        sheet = workbook.create_sheet("Commits")
        self._append_header(sheet, ["Codice", "Descrizione"])

        for row in rows:
            sheet.append([
                row.code,
                row.description,
            ])

        self._autosize_columns(sheet)

    def _write_trips(self, workbook: Workbook, rows: list[TripExportRow]) -> None:
        sheet = workbook.create_sheet("Trips")
        self._append_header(sheet, [
            "Utente",
            "Auto",
            "Commits",
            "Partenza",
            "Arrivo",
            "Data inizio",
            "Data fine",
            "Km iniziali",
            "Km finali",
            "Distanza",
            "Durata minuti",
            "Stato",
        ])

        for row in rows:
            sheet.append([
                row.user_email,
                row.car_plate,
                row.commits,
                row.start_position,
                row.end_position,
                row.start_date,
                row.end_date,
                row.start_km,
                row.end_km,
                row.distance,
                row.duration_minutes,
                row.status,
            ])

        self._autosize_columns(sheet)

    def _write_refuelings(
        self,
        workbook: Workbook,
        rows: list[RefuelingExportRow],
    ) -> None:
        sheet = workbook.create_sheet("Refuelings")
        self._append_header(sheet, [
            "Auto",
            "Carta carburante",
            "Prezzo litro",
            "Litri",
            "Totale",
            "Data",
            "Foto ricevuta",
        ])

        for row in rows:
            sheet.append([
                row.car_plate,
                row.card_number,
                row.liter_price,
                row.liters,
                row.total_price,
                row.date,
                row.receipt_photo,
            ])

        self._autosize_columns(sheet)

    def _append_header(self, sheet: Worksheet, headers: list[str]) -> None:
        sheet.append(headers)

        for cell in sheet[1]:
            cell.style = "Headline 3"

    def _autosize_columns(self, sheet: Worksheet) -> None:
        for column_cells in sheet.columns:
            max_length = 0
            column_letter = column_cells[0].column_letter

            for cell in column_cells:
                if cell.value is not None:
                    max_length = max(max_length, len(str(cell.value)))

            sheet.column_dimensions[column_letter].width = max_length + 2