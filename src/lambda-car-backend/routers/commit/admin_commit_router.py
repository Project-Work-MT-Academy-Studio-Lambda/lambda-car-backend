from uuid import UUID
from io import BytesIO
from zipfile import BadZipFile

from fastapi import APIRouter, Depends, HTTPException, UploadFile, File, status
from openpyxl import load_workbook

from ...commands.commit_commands import (
    CreateCommitCommand,
    UpdateCommitCommand,
    ImportCommitItemCommand,
    ImportCommitsCommand,
)
from ...dependencies import get_commit_service, require_admin
from ...schemas.commit_schemas import (
    CreateCommitRequest,
    UpdateCommitRequest,
    CommitResponse,
    ImportCommitsResponse,
)
from ...services.commit_service import CommitService

from ...logger import get_logger

from ...domain.user import CurrentUser


router = APIRouter(prefix="/admin/commits", tags=["admin-commits"])
logger = get_logger(__name__)

@router.get("/", response_model=list[CommitResponse], status_code=status.HTTP_200_OK)
def list_commits(
    current_user: CurrentUser = Depends(require_admin),
    service: CommitService = Depends(get_commit_service),
):
    logger.debug(f"Admin {current_user.id} is listing all commits")
    commits = service.find_all_commits()
    logger.debug(f"Found {len(commits)} commits in the system")
    return [CommitResponse.from_domain(commit) for commit in commits]

@router.get("/backlog", response_model=list[CommitResponse], status_code=status.HTTP_200_OK)
def list_backlog_commits(
    current_user: CurrentUser = Depends(require_admin),
    service: CommitService = Depends(get_commit_service),
):
    logger.debug(f"Admin {current_user.id} is listing backlog commits")
    try:
        commits = service.find_backlog_commits()
        return [CommitResponse.from_domain(commit) for commit in commits]
    except ValueError as e:
        logger.error(f"Error occurred while fetching commits: {str(e)}")
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail=str(e))

@router.post("/", response_model=CommitResponse, status_code=status.HTTP_201_CREATED)
def create_commit(
    payload: CreateCommitRequest,
    current_user: CurrentUser = Depends(require_admin),
    service: CommitService = Depends(get_commit_service),
):
    try:
        logger.debug(f"Admin {current_user.id} is attempting to create a commit with code: {payload.code}")
        commit = service.create_commit(
            CreateCommitCommand(
                code=payload.code,
                description=payload.description,
            )
        )
        logger.debug(f"Admin {current_user.id} successfully created commit with code: {payload.code}")
        return CommitResponse.from_domain(commit)
    except ValueError as e:
        logger.error(f"Admin {current_user.id} failed to create commit with code: {payload.code}. Error: {e}")
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=str(e))


@router.post(
    "/import-excel",
    response_model=ImportCommitsResponse,
    status_code=status.HTTP_201_CREATED,
)
async def import_commits_excel(
    file: UploadFile = File(...),
    current_user: CurrentUser = Depends(require_admin),
    service: CommitService = Depends(get_commit_service),
):
    filename = file.filename or ""

    if not filename.lower().endswith((".xlsx", ".xlsm")):
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Only .xlsx or .xlsm files are supported",
        )

    try:
        content = await file.read()
        workbook = load_workbook(filename=BytesIO(content), data_only=True)
        sheet = workbook.active

        items: list[ImportCommitItemCommand] = []

        for row in sheet.iter_rows(min_row=2, values_only=True):
            code = row[0] if len(row) > 0 else None
            description = row[1] if len(row) > 1 else None

            if code is None and description is None:
                continue

            items.append(
                ImportCommitItemCommand(
                    code=str(code or "").strip(),
                    description=str(description or "").strip(),
                )
            )

        result = service.import_commits(ImportCommitsCommand(items=items))
        logger.debug(
            f"Admin {current_user.id} imported commits from Excel: "
            f"{result['created']} created, {result['updated']} updated, {result['skipped']} skipped"
        )
        return ImportCommitsResponse(**result)

    except ValueError as e:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=str(e))
    except BadZipFile:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Invalid Excel file")


@router.get("/{commit_id}", response_model=CommitResponse)
def get_commit(
    commit_id: UUID,
    current_user: CurrentUser = Depends(require_admin),
    service: CommitService = Depends(get_commit_service),
):
    logger.debug(f"Admin {current_user.id} is attempting to retrieve commit with ID: {commit_id}")
    try:
        commit = service.get_commit(commit_id)
        logger.debug(f"Admin {current_user.id} successfully retrieved commit with ID: {commit_id}")
        return CommitResponse.from_domain(commit)
    except ValueError as e:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail=str(e))


@router.put("/{commit_id}", response_model=CommitResponse)
def update_commit(
    commit_id: UUID,
    payload: UpdateCommitRequest,
    current_user: CurrentUser = Depends(require_admin),
    service: CommitService = Depends(get_commit_service),
):
    logger.debug(f"Admin {current_user.id} is attempting to update commit with ID: {commit_id}")
    try:
        commit = service.update_commit(
            UpdateCommitCommand(
                commit_id=commit_id,
                code=payload.code,
                description=payload.description,
            )
        )
        logger.debug(f"Admin {current_user.id} successfully updated commit with ID: {commit_id}")
        return CommitResponse.from_domain(commit)
    except ValueError as e:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=str(e))


@router.delete("/{commit_id}", status_code=status.HTTP_204_NO_CONTENT)
def delete_commit(
    commit_id: UUID,
    current_user: CurrentUser = Depends(require_admin),
    service: CommitService = Depends(get_commit_service),
):
    logger.debug(f"Admin {current_user.id} is attempting to delete commit with ID: {commit_id}")
    try:
        service.delete_commit(commit_id)
        logger.debug(f"Admin {current_user.id} successfully deleted commit with ID: {commit_id}")
    except ValueError as e:
        logger.error(f"Admin {current_user.id} failed to delete commit with ID: {commit_id}. Error: {e}")
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail=str(e))
